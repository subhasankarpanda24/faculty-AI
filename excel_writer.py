"""
excel_writer.py — Write faculty data to a richly formatted Excel workbook
=========================================================================
Creates output/NIST_Faculty_Directory.xlsx with three sheets:

  Sheet 1 "All Faculty"       — Full faculty table, sorted by last name,
                                 yellow N/A cells, Profile Completeness %,
                                 Has PhD column, auto-widths, frozen header.
  Sheet 2 "Department Summary" — Faculty count per department + bar chart.
  Sheet 3 "Incomplete Profiles"— Only rows with >5 missing fields.

Dependencies: openpyxl, (data supplied as list of dicts)
"""

import os
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════

# Column order for the "All Faculty" sheet
COLUMNS = [
    "name", "title", "first_name", "last_name", "designation",
    "department", "subjects", "research_areas", "qualification",
    "experience", "room_no", "available_days", "available_time",
    "consultation_mode", "email", "phone", "profile_url",
    "photo_url", "bio", "has_phd", "profile_completeness"
]

# Pretty header names for display
COLUMN_HEADERS = {
    "name": "Full Name",
    "title": "Title",
    "first_name": "First Name",
    "last_name": "Last Name",
    "designation": "Designation",
    "department": "Department",
    "subjects": "Core Subjects",
    "research_areas": "Research Areas",
    "qualification": "Qualification",
    "experience": "Years of Experience",
    "room_no": "Room / Cabin No.",
    "available_days": "Available Days",
    "available_time": "Available Timings",
    "consultation_mode": "Consultation Mode",
    "email": "Email",
    "phone": "Phone",
    "profile_url": "Profile URL",
    "photo_url": "Photo URL",
    "bio": "Short Bio",
    "has_phd": "Has PhD",
    "profile_completeness": "Profile Completeness %",
}

# Fields that count towards "missing" for Incomplete Profiles sheet
COMPLETENESS_FIELDS = [
    "designation", "department", "subjects", "research_areas",
    "qualification", "experience", "room_no", "available_days",
    "available_time", "consultation_mode", "email", "phone", "bio"
]

# Styling
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


# ═══════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def _count_missing(record: Dict) -> int:
    """Count the number of N/A or empty fields in a record."""
    return sum(
        1 for field in COMPLETENESS_FIELDS
        if str(record.get(field, "N/A")).strip().upper() in ("N/A", "", "NONE")
    )


def _auto_width(ws, min_width=10, max_width=45):
    """Auto-adjust column widths based on content length."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_len = len(str(cell.value))
                    max_len = max(max_len, min(cell_len + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# ═══════════════════════════════════════════════════════════════
# SHEET 1: ALL FACULTY
# ═══════════════════════════════════════════════════════════════

def _write_all_faculty(wb: Workbook, records: List[Dict]):
    """Write Sheet 1 — All Faculty with full formatting."""
    ws = wb.active
    ws.title = "All Faculty"

    # Sort by last name alphabetically
    sorted_records = sorted(records, key=lambda r: (r.get("last_name", "") or "").lower())

    # Write headers
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_HEADERS.get(col_key, col_key))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, record in enumerate(sorted_records, start=2):
        for col_idx, col_key in enumerate(COLUMNS, start=1):
            value = record.get(col_key, "N/A")

            # Convert booleans to readable strings
            if isinstance(value, bool):
                value = "Yes" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

            # Highlight N/A cells in yellow
            str_val = str(value).strip().upper()
            if str_val in ("N/A", "", "NONE"):
                cell.fill = YELLOW_FILL

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Auto-adjust column widths
    _auto_width(ws)

    print(f"[ExcelWriter] Sheet 'All Faculty' written: {len(sorted_records)} rows")


# ═══════════════════════════════════════════════════════════════
# SHEET 2: DEPARTMENT SUMMARY
# ═══════════════════════════════════════════════════════════════

def _write_department_summary(wb: Workbook, records: List[Dict]):
    """Write Sheet 2 — Department summary with bar chart."""
    ws = wb.create_sheet("Department Summary")

    # Count faculty per department
    dept_counts = {}
    for r in records:
        dept = r.get("department", "N/A")
        if dept in ("N/A", "", None):
            dept = "Unassigned"
        dept_counts[dept] = dept_counts.get(dept, 0) + 1

    # Sort alphabetically
    sorted_depts = sorted(dept_counts.items(), key=lambda x: x[0])

    # Write headers
    headers = ["Department", "Faculty Count"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Write data
    for row_idx, (dept, count) in enumerate(sorted_depts, start=2):
        ws.cell(row=row_idx, column=1, value=dept).font = DATA_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=count).font = DATA_FONT
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")

    # Add bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Faculty Count by Department"
    chart.y_axis.title = "Number of Faculty"
    chart.x_axis.title = "Department"
    chart.style = 10

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(sorted_depts) + 1)
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_depts) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.shape = 4
    chart.width = 25
    chart.height = 15

    # Place chart below the table
    chart_cell = f"D2"
    ws.add_chart(chart, chart_cell)

    # Auto-adjust widths
    _auto_width(ws)

    print(f"[ExcelWriter] Sheet 'Department Summary' written: {len(sorted_depts)} departments")


# ═══════════════════════════════════════════════════════════════
# SHEET 3: INCOMPLETE PROFILES
# ═══════════════════════════════════════════════════════════════

def _write_incomplete_profiles(wb: Workbook, records: List[Dict]):
    """Write Sheet 3 — Only rows with more than 5 missing fields."""
    ws = wb.create_sheet("Incomplete Profiles")

    # Filter records with >5 missing fields
    incomplete = [r for r in records if _count_missing(r) > 5]
    # Sort by most missing first
    incomplete.sort(key=lambda r: _count_missing(r), reverse=True)

    # Add "Missing Fields Count" as an extra column
    display_columns = COLUMNS + ["missing_count"]
    display_headers = dict(COLUMN_HEADERS)
    display_headers["missing_count"] = "Missing Fields Count"

    # Write headers
    for col_idx, col_key in enumerate(display_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display_headers.get(col_key, col_key))
        cell.fill = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, record in enumerate(incomplete, start=2):
        for col_idx, col_key in enumerate(display_columns, start=1):
            if col_key == "missing_count":
                value = _count_missing(record)
            else:
                value = record.get(col_key, "N/A")

            if isinstance(value, bool):
                value = "Yes" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

            str_val = str(value).strip().upper()
            if str_val in ("N/A", "", "NONE"):
                cell.fill = YELLOW_FILL

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-adjust widths
    _auto_width(ws)

    print(f"[ExcelWriter] Sheet 'Incomplete Profiles' written: {len(incomplete)} rows (of {len(records)} total)")


# ═══════════════════════════════════════════════════════════════
# MAIN EXPORT FUNCTION
# ═══════════════════════════════════════════════════════════════

def write_faculty_excel(records: List[Dict], output_path: str = None) -> str:
    """
    Write all faculty records to a formatted Excel workbook.

    Args:
        records: List of faculty dicts (from scraper.py)
        output_path: Optional path; defaults to output/NIST_Faculty_Directory.xlsx

    Returns:
        The absolute path of the written file.
    """
    if output_path is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(base_dir, "output")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "NIST_Faculty_Directory.xlsx")

    print(f"\n[ExcelWriter] Writing {len(records)} faculty records to Excel...")

    wb = Workbook()

    # Sheet 1: All Faculty
    _write_all_faculty(wb, records)

    # Sheet 2: Department Summary
    _write_department_summary(wb, records)

    # Sheet 3: Incomplete Profiles
    _write_incomplete_profiles(wb, records)

    # Save
    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    print(f"[ExcelWriter] ✅ Excel saved: {abs_path}")

    return abs_path


# ═══════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Quick test: generate sample records and write
    from scraper import scrape_nist_faculty
    records = scrape_nist_faculty()
    write_faculty_excel(records)
